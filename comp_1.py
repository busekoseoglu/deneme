# %% [RAPOR] - MODEL SONUCUNDAN ÇOK SHEET'Lİ EXCEL RAPORU
#
# Bu hücre SOLVE hücresinden sonra çalıştırılır.
#
# Güncel model yapısı:
# - Tek izin/OFF kaynağı: tip_map
# - tip_map tipleri: izin, off_t, off_t2
# - x ve work solver kararlarından okunur
# - pair_off varsa doğrudan solver'dan okunur
# - aylik_mesai_gun varsa aylık mesai toplamı solver'dan okunur
#
# TELAFİ NOTU:
# Güncel modelde gün bazlı "telafi" karar değişkeni yoktur.
# Bu rapor telafi gününü şu şekilde sınıflandırır:
# 1) Bir haftadaki off_t + off_t2 talebi 2'yi aşarsa borç oluşur.
# 2) Başka bir haftada haftalık normal hedefin üzerindeki çalışma,
#    önce açık OFF borcuna atanır.
# 3) Kalan ekstra çalışma, aylik_mesai_gun toplamıyla uyumlu olacak
#    şekilde MESAİ olarak işaretlenir.
#
# Böylece raporda telafi borcu ve ödeme haftası görünür.
# Ancak gün bazlı telafinin solver tarafından kesin seçilmesini istiyorsan
# ileride modele ayrıca telafi_gun karar değişkeni eklenmelidir.

from collections import defaultdict, deque
from datetime import datetime, date, timedelta

import pandas as pd
from ortools.sat.python import cp_model
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter


# ------------------------------------------------------------
# 0) TEMEL KONTROLLER
# ------------------------------------------------------------

if "status" not in globals():
    raise NameError("Önce modeli solve et. 'status' bulunamadı.")

if status not in (cp_model.FEASIBLE, cp_model.OPTIMAL):
    raise ValueError(
        f"Excel yalnızca FEASIBLE/OPTIMAL çözümden üretilebilir. "
        f"Mevcut status: {solver.StatusName(status)}"
    )

zorunlu_degiskenler = [
    "solver",
    "x",
    "work",
    "tip_map",
    "AGENTS",
    "PLAN_GUNLER",
    "gun_vardiyalari",
    "df_tam",
]

eksik_degiskenler = [
    isim for isim in zorunlu_degiskenler
    if isim not in globals()
]

if eksik_degiskenler:
    raise NameError(
        "Eksik değişkenler: " + ", ".join(eksik_degiskenler)
    )


# ------------------------------------------------------------
# 1) AYARLAR
# ------------------------------------------------------------

EXCEL_DOSYA_ADI = globals().get(
    "EXCEL_DOSYA_ADI",
    "vardiya_planlama_model_raporu.xlsx"
)

RAPOR_YOLU = str(Path.cwd() / EXCEL_DOSYA_ADI)

RENKLER = {
    "BASLIK": "#1F4E78",
    "ALT_BASLIK": "#D9EAF7",
    "NORM": "#E2F0D9",
    "MESAI": "#F8CBAD",
    "TELAFI": "#FFD966",
    "IZIN": "#D9E1F2",
    "OFF_T": "#B4C6E7",
    "OFF_T2": "#8EA9DB",
    "NORMAL_OFF": "#E7E6E6",
    "RESMI_TATIL": "#F4B183",
    "ARIFE": "#FFE699",
    "CIFT_OFF": "#C6E0B4",
    "IHLAL": "#F4CCCC",
    "OK": "#D9EAD3",
    "BEYAZ": "#FFFFFF",
    "YAZI": "#1F1F1F",
}

PLAN_TARIHLER = sorted({
    pd.to_datetime(ds).date()
    for ds in PLAN_GUNLER
})

PLAN_TARIH_SET = set(PLAN_TARIHLER)

DATE_TO_DS = {
    pd.to_datetime(ds).date(): ds
    for ds in PLAN_GUNLER
}

RESMI_TATIL_SET = set()

if "resmi_tatil_date_set" in globals():
    RESMI_TATIL_SET = {
        pd.to_datetime(d).date()
        for d in resmi_tatil_date_set
    }
elif "resmi_tatil_plan_gunleri" in globals():
    RESMI_TATIL_SET = {
        pd.to_datetime(d).date()
        for d in resmi_tatil_plan_gunleri
    }
elif "RESMI_TATIL_GUNLERI" in globals():
    RESMI_TATIL_SET = {
        pd.to_datetime(d).date()
        for d in RESMI_TATIL_GUNLERI
    }

ARIFE_SET = set()

if "arife_plan_gunleri" in globals():
    ARIFE_SET = {
        pd.to_datetime(d).date()
        for d in arife_plan_gunleri
    }
elif "ARIFE_GUNLERI" in globals():
    for d in ARIFE_GUNLERI:
        try:
            ARIFE_SET.add(pd.to_datetime(d).date())
        except Exception:
            pass


# ------------------------------------------------------------
# 2) YARDIMCI FONKSİYONLAR
# ------------------------------------------------------------

def norm_agent(value):
    return str(value).strip()


def tip_getir(agent, tarih):
    tip = tip_map.get(norm_agent(agent), {}).get(tarih)
    return str(tip).strip().lower() if tip is not None else ""


def hafta_key(tarih):
    iso = tarih.isocalendar()
    return f"{iso.year}-W{iso.week:02d}"


def gun_adi_tr(tarih):
    gunler = [
        "Pzt", "Sal", "Çar", "Per", "Cum", "Cmt", "Paz"
    ]
    return gunler[tarih.weekday()]


def vardiya_str(v):
    if isinstance(v, tuple) and len(v) == 2:
        return f"{v[0]}-{v[1]}"
    return str(v)


def saat_formatla(value):
    if value is None or value == "":
        return ""

    if hasattr(value, "strftime"):
        try:
            return value.strftime("%H:%M")
        except Exception:
            pass

    text = str(value).strip()

    if " " in text:
        text = text.split(" ")[-1]

    return text[:5]


def vardiya_detay_str(ds, v):
    vardiya_kodu = vardiya_str(v)

    if "saat" not in globals():
        return vardiya_kodu

    vardiya_saati = saat.get((ds, v))

    if not vardiya_saati:
        return vardiya_kodu

    try:
        baslangic, bitis = vardiya_saati
    except Exception:
        return vardiya_kodu

    baslangic = saat_formatla(baslangic)
    bitis = saat_formatla(bitis)

    if baslangic and bitis:
        return f"{vardiya_kodu} | {baslangic}-{bitis}"

    if baslangic:
        return f"{vardiya_kodu} | {baslangic}"

    return vardiya_kodu


def agent_meta_map_olustur():
    sonuc = {}

    df_local = df_tam.copy()
    df_local["agent_user_code"] = (
        df_local["agent_user_code"]
        .astype(str)
        .str.strip()
    )

    for _, row in df_local.iterrows():
        a = row["agent_user_code"]

        sonuc[a] = {
            "agent_name": row.get("agent_name", ""),
            "teamleader_name": row.get("teamleader_name", ""),
            "takim": (
                row.get("takim", "")
                or row.get("team", "")
                or row.get("team_name", "")
            ),
            "working_main_group": row.get("working_main_group", ""),
            "mesaiye_kalamaz_flg": row.get(
                "mesaiye_kalamaz_flg", 0
            ),
            "hamile_flg": row.get("hamile_flg", 0),
            "sut_izni_flg": row.get("sut_izni_flg", 0),
            "dogum_izni_flg": row.get("dogum_izni_flg", 0),
            "idari_izinli_flg": row.get("idari_izinli_flg", 0),
        }

    return sonuc


AGENT_META = agent_meta_map_olustur()


# ------------------------------------------------------------
# 3) SOLVER ATAMALARINI OKU
# ------------------------------------------------------------

atama_map = {}
work_map = {}

for a_raw in AGENTS:
    a = norm_agent(a_raw)

    for tarih in PLAN_TARIHLER:
        ds = DATE_TO_DS[tarih]

        work_value = 0
        if (a, ds) in work:
            work_value = int(solver.Value(work[(a, ds)]))

        work_map[(a, tarih)] = work_value

        secilen_vardiyalar = []

        for v in gun_vardiyalari.get(ds, []):
            if (
                (a, ds, v) in x
                and solver.Value(x[(a, ds, v)]) == 1
            ):
                secilen_vardiyalar.append(v)

        atama_map[(a, tarih)] = secilen_vardiyalar


# ------------------------------------------------------------
# 4) HAFTALIK BAZ NORMAL HEDEF
# ------------------------------------------------------------

hafta_tarihleri = defaultdict(list)

for tarih in PLAN_TARIHLER:
    hafta_tarihleri[hafta_key(tarih)].append(tarih)


def haftalik_baz_hedef(agent, wk):
    tarihler = hafta_tarihleri[wk]

    hafta_ici_is_gunleri = [
        t for t in tarihler
        if t.weekday() < 5 and t not in RESMI_TATIL_SET
    ]

    izin_is_gunu = sum(
        1 for t in hafta_ici_is_gunleri
        if tip_getir(agent, t) == "izin"
    )

    return max(
        len(hafta_ici_is_gunleri) - izin_is_gunu,
        0
    )


# ------------------------------------------------------------
# 5) HAFTALIK GERÇEK OFF, OFF BORCU VE TELAFİ
# ------------------------------------------------------------
#
# Doğru raporlama mantığı:
# - İzin ve resmî tatil gerçek OFF sayılmaz.
# - off_t / off_t2 gerçek OFF sayılır.
# - Bunların dışındaki work=0 günleri gerçek OFF sayılır.
# - Haftada 2 gerçek OFF varsa hafta sonu çalışmaları NORM'dur.
# - Haftada 1 gerçek OFF varsa 6. çalışma günlerinden biri TELAFİ'dir.
# - TELAFİ yalnızca başka haftadaki OFF borcuna karşılık atanır.

haftalik_gercek_off = {}
haftalik_calisma_gunleri = {}
haftalik_hafta_sonu_calisma = {}

for a_raw in AGENTS:
    a = norm_agent(a_raw)

    for wk, tarihler_raw in hafta_tarihleri.items():
        tarihler = sorted(tarihler_raw)
        gercek_off_tarihleri = []
        calisma_tarihleri = []
        hafta_sonu_calisma_tarihleri = []

        for tarih in tarihler:
            tip = tip_getir(a, tarih)
            calisti = work_map.get((a, tarih), 0) == 1

            if calisti:
                calisma_tarihleri.append(tarih)
                if tarih.weekday() >= 5:
                    hafta_sonu_calisma_tarihleri.append(tarih)

            if tip == "izin":
                continue

            if tarih in RESMI_TATIL_SET:
                continue

            if tip in {"off_t", "off_t2"}:
                gercek_off_tarihleri.append(tarih)
            elif not calisti:
                gercek_off_tarihleri.append(tarih)

        haftalik_gercek_off[(a, wk)] = gercek_off_tarihleri
        haftalik_calisma_gunleri[(a, wk)] = calisma_tarihleri
        haftalik_hafta_sonu_calisma[(a, wk)] = hafta_sonu_calisma_tarihleri


off_borc_rows = []
telafi_gunleri = {}
telafi_detay_rows = []
acik_borc_agent = {}

for a_raw in AGENTS:
    a = norm_agent(a_raw)
    hafta_bilgileri = []

    for wk in sorted(hafta_tarihleri):
        tarihler = sorted(hafta_tarihleri[wk])

        talep_tarihleri = [
            t for t in tarihler
            if tip_getir(a, t) in {"off_t", "off_t2"}
        ]

        talep_sayisi = len(talep_tarihleri)
        borc = max(talep_sayisi - 2, 0)

        gercek_off_tarihleri = haftalik_gercek_off.get((a, wk), [])
        calisma_tarihleri = haftalik_calisma_gunleri.get((a, wk), [])
        hafta_sonu_calisma_tarihleri = haftalik_hafta_sonu_calisma.get((a, wk), [])

        gercek_off_sayisi = len(gercek_off_tarihleri)

        # 2 OFF yerine 1 OFF kullanılmışsa 1 adet telafi ödeme kapasitesi vardır.
        telafi_odeme_kapasitesi = max(2 - gercek_off_sayisi, 0)

        # Telafi günü mümkünse hafta sonu çalışmasından seçilir.
        telafi_aday_gunleri = sorted(
            hafta_sonu_calisma_tarihleri,
            reverse=True,
        )[:telafi_odeme_kapasitesi]

        hafta_bilgileri.append({
            "week": wk,
            "off_request_dates": talep_tarihleri,
            "off_request_count": talep_sayisi,
            "debt": borc,
            "real_off_dates": gercek_off_tarihleri,
            "real_off_count": gercek_off_sayisi,
            "work_dates": calisma_tarihleri,
            "weekend_work_dates": hafta_sonu_calisma_tarihleri,
            "payment_capacity": telafi_odeme_kapasitesi,
            "payment_candidate_dates": telafi_aday_gunleri,
        })

        off_borc_rows.append({
            "agent_user_code": a,
            "hafta": wk,
            "off_talep_sayisi": talep_sayisi,
            "off_talep_tarihleri": ", ".join(t.isoformat() for t in talep_tarihleri),
            "gercek_off_sayisi": gercek_off_sayisi,
            "gercek_off_tarihleri": ", ".join(t.isoformat() for t in gercek_off_tarihleri),
            "standart_off_hakki": 2,
            "telafi_borcu": borc,
            "telafi_odeme_kapasitesi": telafi_odeme_kapasitesi,
            "hafta_sonu_calisma_tarihleri": ", ".join(
                t.isoformat() for t in hafta_sonu_calisma_tarihleri
            ),
        })

    borclar = []

    for info in hafta_bilgileri:
        for _ in range(info["debt"]):
            borclar.append({
                "source_week": info["week"],
                "source_dates": info["off_request_dates"],
                "paid": False,
                "payment_week": "",
                "payment_date": None,
                "payment_type": "",
            })

    kullanilan_odeme_gunleri = set()

    # Önce borçtan sonraki haftalardaki 1-OFF haftalarını kullan.
    for borc in borclar:
        adaylar = []

        for info in hafta_bilgileri:
            if info["week"] <= borc["source_week"]:
                continue

            for tarih in info["payment_candidate_dates"]:
                if tarih not in kullanilan_odeme_gunleri:
                    adaylar.append((info["week"], tarih))

        if adaylar:
            odeme_week, odeme_tarih = sorted(adaylar, key=lambda x: (x[0], x[1]))[0]
            borc["paid"] = True
            borc["payment_week"] = odeme_week
            borc["payment_date"] = odeme_tarih
            borc["payment_type"] = "TELAFİ"
            kullanilan_odeme_gunleri.add(odeme_tarih)
            telafi_gunleri[(a, odeme_tarih)] = {
                "source_week": borc["source_week"],
                "payment_week": odeme_week,
                "payment_type": "TELAFİ",
            }

    # Ay içindeki planlamada ödeme haftası borçtan önceyse ön ödeme olarak eşleştir.
    for borc in borclar:
        if borc["paid"]:
            continue

        adaylar = []

        for info in hafta_bilgileri:
            if info["week"] >= borc["source_week"]:
                continue

            for tarih in info["payment_candidate_dates"]:
                if tarih not in kullanilan_odeme_gunleri:
                    adaylar.append((info["week"], tarih))

        if adaylar:
            odeme_week, odeme_tarih = sorted(
                adaylar,
                key=lambda x: (x[0], x[1]),
                reverse=True,
            )[0]

            borc["paid"] = True
            borc["payment_week"] = odeme_week
            borc["payment_date"] = odeme_tarih
            borc["payment_type"] = "TELAFİ ÖN ÖDEME"
            kullanilan_odeme_gunleri.add(odeme_tarih)
            telafi_gunleri[(a, odeme_tarih)] = {
                "source_week": borc["source_week"],
                "payment_week": odeme_week,
                "payment_type": "TELAFİ ÖN ÖDEME",
            }

    acik_borc = 0

    for borc in borclar:
        if not borc["paid"]:
            acik_borc += 1

        telafi_detay_rows.append({
            "agent_user_code": a,
            "borc_haftasi": borc["source_week"],
            "borca_neden_olan_off_tarihleri": ", ".join(
                t.isoformat() for t in borc["source_dates"]
            ),
            "durum": "ÖDENDİ" if borc["paid"] else "AÇIK BORÇ",
            "odeme_haftasi": borc["payment_week"],
            "odeme_tarihi": (
                borc["payment_date"].isoformat()
                if borc["payment_date"]
                else ""
            ),
            "odeme_tipi": borc["payment_type"],
        })

    acik_borc_agent[a] = acik_borc


# ------------------------------------------------------------
# 6) MESAİ GÜNLERİNİ RAPORLAMA İÇİN DAĞIT
# ------------------------------------------------------------
# aylik_mesai_gun yalnızca toplam mesai gününü verir.
# TELAFİ günleri mesai değildir ve aday havuzundan çıkarılır.

mesai_gunleri = {}

for a_raw in AGENTS:
    a = norm_agent(a_raw)
    mesai_sayisi = 0

    if (
        "aylik_mesai_gun" in globals()
        and isinstance(aylik_mesai_gun, dict)
        and a in aylik_mesai_gun
    ):
        mesai_sayisi = int(solver.Value(aylik_mesai_gun[a]))

    aday_gunler = [
        t for t in PLAN_TARIHLER
        if work_map.get((a, t), 0) == 1
        and (a, t) not in telafi_gunleri
    ]

    aday_gunler = sorted(
        aday_gunler,
        key=lambda t: (
            0 if t in RESMI_TATIL_SET else 1,
            0 if t in ARIFE_SET else 1,
            0 if t.weekday() >= 5 else 1,
            -t.toordinal(),
        ),
    )

    for tarih in aday_gunler[:mesai_sayisi]:
        mesai_gunleri[(a, tarih)] = True


# ------------------------------------------------------------
# 7) GÜNLÜK DURUM SATIRLARI
# ------------------------------------------------------------
# - Hafta içi normal çalışma: sadece vardiya gösterilir, durum etiketi boş kalır.
# - Hafta sonu ve haftada 2 gerçek OFF: NORM.
# - OFF borcunu ödeyen hafta sonu çalışma: TELAFİ.
# - Aylık hedef üzeri çalışma: MESAİ.

gunluk_rows = []
calendar_status = {}

for a_raw in AGENTS:
    a = norm_agent(a_raw)
    meta = AGENT_META.get(a, {})

    for tarih in PLAN_TARIHLER:
        ds = DATE_TO_DS[tarih]
        wk = hafta_key(tarih)
        tip = tip_getir(a, tarih)
        calisti = work_map.get((a, tarih), 0) == 1
        vardiyalar = atama_map.get((a, tarih), [])

        vardiya_text = " / ".join(
            vardiya_detay_str(ds, v)
            for v in vardiyalar
        )

        if tip == "izin":
            durum = "İZİN"
            renk_kodu = "IZIN"

        elif tip == "off_t":
            durum = "OFF_T ✓"
            renk_kodu = "OFF_T"

        elif tip == "off_t2":
            durum = "OFF_T2 ✓"
            renk_kodu = "OFF_T2"

        elif calisti:
            if (a, tarih) in telafi_gunleri:
                durum = "TELAFİ"
                renk_kodu = "TELAFI"

            elif (a, tarih) in mesai_gunleri:
                if tarih in RESMI_TATIL_SET:
                    durum = "RESMİ TATİL MESAİ"
                    renk_kodu = "RESMI_TATIL"
                elif tarih in ARIFE_SET:
                    durum = "ARİFE MESAİ"
                    renk_kodu = "ARIFE"
                else:
                    durum = "MESAİ"
                    renk_kodu = "MESAI"

            elif tarih.weekday() >= 5:
                # Hafta sonu normal çalışma.
                # Örn. Pazartesi ve Pazar OFF, Cumartesi çalışma => NORM.
                durum = "NORM"
                renk_kodu = "NORM"

            else:
                # Hafta içi çalışma için NORM etiketi yazılmaz.
                durum = ""
                renk_kodu = "BEYAZ"

        else:
            if tarih in RESMI_TATIL_SET:
                durum = "RESMİ TATİL"
                renk_kodu = "RESMI_TATIL"
            else:
                durum = "OFF"
                renk_kodu = "NORMAL_OFF"

        calendar_status[(a, tarih)] = {
            "durum": durum,
            "renk_kodu": renk_kodu,
            "vardiya": vardiya_text,
        }

        gunluk_rows.append({
            "agent_user_code": a,
            "agent_name": meta.get("agent_name", ""),
            "takim": meta.get("takim", ""),
            "teamleader_name": meta.get("teamleader_name", ""),
            "tarih": tarih.isoformat(),
            "gun": gun_adi_tr(tarih),
            "hafta": wk,
            "durum": durum,
            "vardiya": vardiya_text,
            "izin_off_tipi": tip,
            "calisti_mi": int(calisti),
            "haftalik_gercek_off_sayisi": len(
                haftalik_gercek_off.get((a, wk), [])
            ),
            "resmi_tatil_mi": int(tarih in RESMI_TATIL_SET),
            "arife_mi": int(tarih in ARIFE_SET),
            "telafi_borc_haftasi": telafi_gunleri.get(
                (a, tarih), {}
            ).get("source_week", ""),
        })


# ------------------------------------------------------------
# 8) ÇİFT OFF SONUÇLARI
# ------------------------------------------------------------

cift_off_rows = []
cift_off_sayisi_agent = defaultdict(int)

pair_off_obj = globals().get("pair_off", None)
weekend_pairs_obj = globals().get("weekend_pairs", [])

pair_off_dict_mi = isinstance(pair_off_obj, dict)
weekend_pairs_gecerli_mi = isinstance(weekend_pairs_obj, (list, tuple))

if weekend_pairs_gecerli_mi:
    for a_raw in AGENTS:
        a = norm_agent(a_raw)

        for pair_index, (sat_ds, sun_ds) in enumerate(weekend_pairs_obj):
            sat_date = pd.to_datetime(sat_ds).date()
            sun_date = pd.to_datetime(sun_ds).date()

            if pair_off_dict_mi:
                var = pair_off_obj.get((a, pair_index))
                pair_value = (
                    int(solver.Value(var))
                    if var is not None
                    else 0
                )
                hesaplama_kaynagi = "solver_pair_off"

            else:
                # pair_off sözlük değilse final takvimden hesapla.
                # İzin gerçek OFF sayılmaz; off_t/off_t2 ve normal work=0
                # gerçek OFF sayılır.
                sat_tip = tip_getir(a, sat_date)
                sun_tip = tip_getir(a, sun_date)

                sat_gercek_off = (
                    sat_tip in {"off_t", "off_t2"}
                    or (
                        sat_tip != "izin"
                        and work_map.get((a, sat_date), 0) == 0
                    )
                )

                sun_gercek_off = (
                    sun_tip in {"off_t", "off_t2"}
                    or (
                        sun_tip != "izin"
                        and work_map.get((a, sun_date), 0) == 0
                    )
                )

                pair_value = int(sat_gercek_off and sun_gercek_off)
                hesaplama_kaynagi = "final_takvimden_hesaplandi"

            if pair_value == 1:
                cift_off_sayisi_agent[a] += 1

            cift_off_rows.append({
                "agent_user_code": a,
                "pair_index": pair_index,
                "cumartesi": sat_date.isoformat(),
                "pazar": sun_date.isoformat(),
                "cumartesi_durum": calendar_status[(a, sat_date)]["durum"],
                "pazar_durum": calendar_status[(a, sun_date)]["durum"],
                "cift_off_mi": pair_value,
                "hesaplama_kaynagi": hesaplama_kaynagi,
            })

print(
    "Çift OFF okuma kaynağı:",
    "solver pair_off sözlüğü" if pair_off_dict_mi
    else "final takvimden hesaplandı"
)


# ------------------------------------------------------------
# 9) AGENT AYLIK ÖZET
# ------------------------------------------------------------

agent_ozet_rows = []

for a_raw in AGENTS:
    a = norm_agent(a_raw)
    meta = AGENT_META.get(a, {})

    durum_sayilari = defaultdict(int)

    for tarih in PLAN_TARIHLER:
        durum = calendar_status[(a, tarih)]["durum"]

        if durum.startswith("RESMİ TATİL MESAİ"):
            durum_sayilari["resmi_tatil_calisma"] += 1
            durum_sayilari["mesai"] += 1
        elif durum.startswith("ARİFE MESAİ"):
            durum_sayilari["arife_calisma"] += 1
            durum_sayilari["mesai"] += 1
        elif durum == "MESAİ":
            durum_sayilari["mesai"] += 1
        elif durum == "TELAFİ":
            durum_sayilari["telafi"] += 1
        elif durum == "NORM":
            durum_sayilari["norm"] += 1
        elif durum == "İZİN":
            durum_sayilari["izin"] += 1
        elif durum == "OFF_T ✓":
            durum_sayilari["off_t"] += 1
        elif durum == "OFF_T2 ✓":
            durum_sayilari["off_t2"] += 1
        elif durum == "OFF":
            durum_sayilari["normal_off"] += 1
        elif durum == "RESMİ TATİL":
            durum_sayilari["resmi_tatil_off"] += 1

    toplam_calisma = sum(
        work_map.get((a, t), 0)
        for t in PLAN_TARIHLER
    )

    agent_ozet_rows.append({
        "agent_user_code": a,
        "agent_name": meta.get("agent_name", ""),
        "takim": meta.get("takim", ""),
        "teamleader_name": meta.get(
            "teamleader_name", ""
        ),
        "working_main_group": meta.get(
            "working_main_group", ""
        ),
        "toplam_calisma_gunu": toplam_calisma,
        "norm_calisma": durum_sayilari["norm"],
        "telafi_calisma": durum_sayilari["telafi"],
        "mesai_calisma": durum_sayilari["mesai"],
        "izin_gunu": durum_sayilari["izin"],
        "off_t_karsilanan": durum_sayilari["off_t"],
        "off_t2_karsilanan": durum_sayilari["off_t2"],
        "normal_off": durum_sayilari["normal_off"],
        "resmi_tatil_calisma": durum_sayilari[
            "resmi_tatil_calisma"
        ],
        "arife_calisma": durum_sayilari[
            "arife_calisma"
        ],
        "cift_off_sayisi": cift_off_sayisi_agent[a],
        "acik_telafi_borcu": acik_borc_agent.get(a, 0),
        "mesaiye_kalamaz_flg": meta.get(
            "mesaiye_kalamaz_flg", 0
        ),
        "hamile_flg": meta.get("hamile_flg", 0),
        "sut_izni_flg": meta.get(
            "sut_izni_flg", 0
        ),
        "dogum_izni_flg": meta.get(
            "dogum_izni_flg", 0
        ),
        "idari_izinli_flg": meta.get(
            "idari_izinli_flg", 0
        ),
    })


# ------------------------------------------------------------
# 10) COVERAGE
# ------------------------------------------------------------

coverage_rows = []

for ds in PLAN_GUNLER:
    tarih = pd.to_datetime(ds).date()

    for v in gun_vardiyalari.get(ds, []):
        atanan = sum(
            int(solver.Value(x[(norm_agent(a), ds, v)]))
            for a in AGENTS
            if (norm_agent(a), ds, v) in x
        )

        required = int(talep.get((ds, v), 0)) \
            if "talep" in globals() else 0

        coverage_rows.append({
            "tarih": tarih.isoformat(),
            "gun": gun_adi_tr(tarih),
            "vardiya": vardiya_str(v),
            "required": required,
            "atanan": atanan,
            "gap": atanan - required,
            "eksik": max(required - atanan, 0),
            "fazla": max(atanan - required, 0),
        })


# ------------------------------------------------------------
# 11) KURAL KONTROLLERİ
# ------------------------------------------------------------

kural_rows = []

# 11.1 Günde en fazla bir vardiya
for a_raw in AGENTS:
    a = norm_agent(a_raw)

    for tarih in PLAN_TARIHLER:
        vardiya_sayisi = len(
            atama_map.get((a, tarih), [])
        )

        if vardiya_sayisi > 1:
            kural_rows.append({
                "kural": "Günde en fazla 1 vardiya",
                "agent_user_code": a,
                "tarih_hafta": tarih.isoformat(),
                "sonuc": "İHLAL",
                "detay": f"{vardiya_sayisi} vardiya",
            })

# 11.2 Maksimum 6 gün üst üste
for a_raw in AGENTS:
    a = norm_agent(a_raw)
    streak = 0
    streak_start = None

    for tarih in PLAN_TARIHLER:
        if work_map.get((a, tarih), 0) == 1:
            if streak == 0:
                streak_start = tarih
            streak += 1

            if streak > 6:
                kural_rows.append({
                    "kural": "Maksimum 6 gün üst üste",
                    "agent_user_code": a,
                    "tarih_hafta": (
                        f"{streak_start.isoformat()} - "
                        f"{tarih.isoformat()}"
                    ),
                    "sonuc": "İHLAL",
                    "detay": f"{streak} gün üst üste",
                })
        else:
            streak = 0
            streak_start = None

# 11.3 İzin/OFF tarihinde çalışma
for a_raw in AGENTS:
    a = norm_agent(a_raw)

    for tarih in PLAN_TARIHLER:
        tip = tip_getir(a, tarih)

        if (
            tip in {"izin", "off_t", "off_t2"}
            and work_map.get((a, tarih), 0) == 1
        ):
            kural_rows.append({
                "kural": "İzin/OFF tarihinde çalışmama",
                "agent_user_code": a,
                "tarih_hafta": tarih.isoformat(),
                "sonuc": "İHLAL",
                "detay": tip,
            })

# 11.4 OFF talebi karşılanma
for a_raw in AGENTS:
    a = norm_agent(a_raw)

    for tarih, tip in tip_map.get(a, {}).items():
        tip_str = str(tip).strip().lower()

        if (
            tarih in PLAN_TARIH_SET
            and tip_str in {"off_t", "off_t2"}
        ):
            karsilandi = (
                work_map.get((a, tarih), 0) == 0
            )

            kural_rows.append({
                "kural": "OFF talebi karşılanma",
                "agent_user_code": a,
                "tarih_hafta": tarih.isoformat(),
                "sonuc": (
                    "OK" if karsilandi else "İHLAL"
                ),
                "detay": tip_str,
            })

# 11.5 Çift OFF
for a_raw in AGENTS:
    a = norm_agent(a_raw)

    tum_plan_izinli = all(
        tip_getir(a, t) == "izin"
        for t in PLAN_TARIHLER
    )

    if not tum_plan_izinli:
        cift_sayi = cift_off_sayisi_agent.get(a, 0)

        kural_rows.append({
            "kural": "Ayda en az 1 çift OFF",
            "agent_user_code": a,
            "tarih_hafta": "",
            "sonuc": (
                "OK" if cift_sayi >= 1 else "İHLAL"
            ),
            "detay": f"çift OFF sayısı: {cift_sayi}",
        })

# 11.6 Telafi borcu
for a_raw in AGENTS:
    a = norm_agent(a_raw)
    acik = acik_borc_agent.get(a, 0)

    kural_rows.append({
        "kural": "OFF telafi borcu",
        "agent_user_code": a,
        "tarih_hafta": "",
        "sonuc": "OK" if acik == 0 else "İHLAL",
        "detay": f"açık borç: {acik}",
    })

# 11.7 Aylık çalışma hedefi
if "aylik_calisma_debug_df" in globals():
    debug_map = {}

    for _, row in aylik_calisma_debug_df.iterrows():
        debug_map[norm_agent(
            row["agent_user_code"]
        )] = row.to_dict()

    for a_raw in AGENTS:
        a = norm_agent(a_raw)

        toplam_calisma = sum(
            work_map.get((a, t), 0)
            for t in PLAN_TARIHLER
            if t not in RESMI_TATIL_SET
        )

        info = debug_map.get(a, {})
        min_calisma = int(
            info.get("min_calisma", 0)
        )
        max_calisma = int(
            info.get(
                "max_calisma",
                min_calisma
            )
        )

        uygun = (
            min_calisma
            <= toplam_calisma
            <= max_calisma
        )

        kural_rows.append({
            "kural": "Aylık çalışma hedefi",
            "agent_user_code": a,
            "tarih_hafta": "",
            "sonuc": "OK" if uygun else "İHLAL",
            "detay": (
                f"gerçekleşen={toplam_calisma}, "
                f"min={min_calisma}, max={max_calisma}"
            ),
        })

# Hiç ihlal oluşmayan bazı kontrollerin de görünmesi için
# özet satırları
kural_ozet_rows = []

for kural in sorted({
    row["kural"] for row in kural_rows
}):
    satirlar = [
        row for row in kural_rows
        if row["kural"] == kural
    ]
    ihlal = sum(
        1 for row in satirlar
        if row["sonuc"] == "İHLAL"
    )

    kural_ozet_rows.append({
        "kural": kural,
        "kontrol_kaydi": len(satirlar),
        "ihlal_sayisi": ihlal,
        "genel_sonuc": "OK" if ihlal == 0 else "İHLAL",
    })


# ------------------------------------------------------------
# 12) OPENPYXL YARDIMCILARI
# ------------------------------------------------------------

ince_kenarlik = Side(style="thin", color="D9E1F2")
standart_border = Border(
    left=ince_kenarlik,
    right=ince_kenarlik,
    top=ince_kenarlik,
    bottom=ince_kenarlik,
)


def hex_renk(renk):
    return str(renk).replace("#", "").upper()


def fill_olustur(renk):
    return PatternFill(
        fill_type="solid",
        fgColor=hex_renk(renk),
    )


def sheet_yaz(ws, headers, rows):
    ws.append(headers)

    for row in rows:
        ws.append([
            row.get(header, "")
            for header in headers
        ])

    for cell in ws[1]:
        cell.fill = fill_olustur(RENKLER["BASLIK"])
        cell.font = Font(
            bold=True,
            color=hex_renk(RENKLER["BEYAZ"]),
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = standart_border

    for row in ws.iter_rows(
        min_row=2,
        max_row=ws.max_row,
        min_col=1,
        max_col=ws.max_column,
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True,
            )
            cell.border = standart_border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    kolon_genisliklerini_ayarla(ws)


def kolon_genisliklerini_ayarla(ws, max_genislik=35):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)

        for cell in ws[col_letter]:
            value = "" if cell.value is None else str(cell.value)
            satir_uzunlugu = max(
                [len(parca) for parca in value.split("\n")]
                or [0]
            )
            max_len = max(max_len, satir_uzunlugu)

        ws.column_dimensions[col_letter].width = min(
            max(max_len + 2, 10),
            max_genislik,
        )


def bos_default_sheet_sil(wb):
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]


# ------------------------------------------------------------
# 13) WORKBOOK OLUŞTUR
# ------------------------------------------------------------

wb = Workbook()


# 13.1 Renk Açıklamaları
ws = wb.active
ws.title = "Renk Açıklamaları"

legend_rows = [
    ["Durum", "Anlam"],
    ["NORM", "Normal çalışma"],
    ["MESAİ", "Aylık hedef üzerindeki çalışma"],
    ["TELAFİ", "2'yi aşan haftalık OFF talebinin başka haftada ödenmesi"],
    ["İZİN", "Gerçek izin"],
    ["OFF_T ✓", "Agent OFF_T talebi karşılandı"],
    ["OFF_T2 ✓", "Agent OFF_T2 talebi karşılandı"],
    ["OFF", "Modelin verdiği normal OFF"],
    ["RESMİ TATİL", "Resmî tatilde çalışmadı"],
    ["RESMİ TATİL MESAİ", "Resmî tatil çalışması"],
    ["ARİFE MESAİ", "Arife çalışması"],
]

for row in legend_rows:
    ws.append(row)

for cell in ws[1]:
    cell.fill = fill_olustur(RENKLER["BASLIK"])
    cell.font = Font(
        bold=True,
        color=hex_renk(RENKLER["BEYAZ"]),
    )

legend_renk_map = {
    "NORM": "NORM",
    "MESAİ": "MESAI",
    "TELAFİ": "TELAFI",
    "İZİN": "IZIN",
    "OFF_T ✓": "OFF_T",
    "OFF_T2 ✓": "OFF_T2",
    "OFF": "NORMAL_OFF",
    "RESMİ TATİL": "RESMI_TATIL",
    "RESMİ TATİL MESAİ": "RESMI_TATIL",
    "ARİFE MESAİ": "ARIFE",
}

for row_no in range(2, len(legend_rows) + 1):
    durum = ws.cell(row=row_no, column=1).value
    renk_anahtari = legend_renk_map.get(durum)

    if renk_anahtari:
        for col_no in [1, 2]:
            ws.cell(
                row=row_no,
                column=col_no,
            ).fill = fill_olustur(RENKLER[renk_anahtari])

kolon_genisliklerini_ayarla(ws)
ws.freeze_panes = "A2"


# 13.2 Agent Takvimi
ws = wb.create_sheet("Agent Takvimi")

sabit_headers = [
    "agent_user_code",
    "agent_name",
    "takim",
    "teamleader_name",
]

tarih_headers = [
    f"{t.day:02d} {gun_adi_tr(t)}"
    for t in PLAN_TARIHLER
]

takvim_headers = sabit_headers + tarih_headers
ws.append(takvim_headers)

for a_raw in AGENTS:
    a = norm_agent(a_raw)
    meta = AGENT_META.get(a, {})

    row = [
        a,
        meta.get("agent_name", ""),
        meta.get("takim", ""),
        meta.get("teamleader_name", ""),
    ]

    for tarih in PLAN_TARIHLER:
        info = calendar_status[(a, tarih)]
        cell_text = info["durum"]

        if info["vardiya"]:
            cell_text += f"\n{info['vardiya']}"

        if (a, tarih) in telafi_gunleri:
            cell_text += (
                "\nBorç: "
                + telafi_gunleri[(a, tarih)]["source_week"]
            )

        row.append(cell_text)

    ws.append(row)

for cell in ws[1]:
    cell.fill = fill_olustur(RENKLER["BASLIK"])
    cell.font = Font(
        bold=True,
        color=hex_renk(RENKLER["BEYAZ"]),
    )
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

ws.freeze_panes = "E2"
ws.auto_filter.ref = ws.dimensions

ws.column_dimensions["A"].width = 17
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 22

for tarih_index, tarih in enumerate(PLAN_TARIHLER, start=5):
    ws.column_dimensions[
        get_column_letter(tarih_index)
    ].width = 16

for row_index, a_raw in enumerate(AGENTS, start=2):
    a = norm_agent(a_raw)

    for tarih_index, tarih in enumerate(PLAN_TARIHLER, start=5):
        cell = ws.cell(
            row=row_index,
            column=tarih_index,
        )
        info = calendar_status[(a, tarih)]
        cell.fill = fill_olustur(
            RENKLER[info["renk_kodu"]]
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = standart_border

for row in ws.iter_rows(
    min_row=2,
    max_row=ws.max_row,
    min_col=1,
    max_col=4,
):
    for cell in row:
        cell.border = standart_border
        cell.alignment = Alignment(
            vertical="center",
            wrap_text=True,
        )


# 13.3 Agent Özet
ws = wb.create_sheet("Agent Özet")

agent_ozet_headers = [
    "agent_user_code",
    "agent_name",
    "takim",
    "teamleader_name",
    "working_main_group",
    "toplam_calisma_gunu",
    "norm_calisma",
    "telafi_calisma",
    "mesai_calisma",
    "izin_gunu",
    "off_t_karsilanan",
    "off_t2_karsilanan",
    "normal_off",
    "resmi_tatil_calisma",
    "arife_calisma",
    "cift_off_sayisi",
    "acik_telafi_borcu",
    "mesaiye_kalamaz_flg",
    "hamile_flg",
    "sut_izni_flg",
    "dogum_izni_flg",
    "idari_izinli_flg",
]

sheet_yaz(ws, agent_ozet_headers, agent_ozet_rows)

if ws.max_row >= 2:
    ws.conditional_formatting.add(
        f"Q2:Q{ws.max_row}",
        CellIsRule(
            operator="greaterThan",
            formula=["0"],
            fill=fill_olustur(RENKLER["IHLAL"]),
        ),
    )


# 13.4 Günlük Atamalar
ws = wb.create_sheet("Günlük Atamalar")

gunluk_headers = [
    "agent_user_code",
    "agent_name",
    "takim",
    "teamleader_name",
    "tarih",
    "gun",
    "hafta",
    "durum",
    "vardiya",
    "izin_off_tipi",
    "calisti_mi",
    "resmi_tatil_mi",
    "arife_mi",
    "telafi_borc_haftasi",
]

sheet_yaz(ws, gunluk_headers, gunluk_rows)


# 13.5 OFF Talepleri
ws = wb.create_sheet("OFF Talepleri")

off_talep_rows = []

for a_raw in AGENTS:
    a = norm_agent(a_raw)

    for tarih, tip in sorted(tip_map.get(a, {}).items()):
        tip_str = str(tip).strip().lower()

        if (
            tarih not in PLAN_TARIH_SET
            or tip_str not in {"off_t", "off_t2"}
        ):
            continue

        karsilandi = work_map.get((a, tarih), 0) == 0

        off_talep_rows.append({
            "agent_user_code": a,
            "agent_name": AGENT_META.get(a, {}).get(
                "agent_name", ""
            ),
            "tarih": tarih.isoformat(),
            "hafta": hafta_key(tarih),
            "talep_tipi": tip_str,
            "karsilandi_mi": "EVET" if karsilandi else "HAYIR",
            "takvim_durumu": calendar_status[(a, tarih)]["durum"],
        })

sheet_yaz(
    ws,
    [
        "agent_user_code",
        "agent_name",
        "tarih",
        "hafta",
        "talep_tipi",
        "karsilandi_mi",
        "takvim_durumu",
    ],
    off_talep_rows,
)

if ws.max_row >= 2:
    ws.conditional_formatting.add(
        f"F2:F{ws.max_row}",
        FormulaRule(
            formula=['F2="HAYIR"'],
            fill=fill_olustur(RENKLER["IHLAL"]),
        ),
    )
    ws.conditional_formatting.add(
        f"F2:F{ws.max_row}",
        FormulaRule(
            formula=['F2="EVET"'],
            fill=fill_olustur(RENKLER["OK"]),
        ),
    )


# 13.6 Telafi Takibi
ws = wb.create_sheet("Telafi Takibi")

sheet_yaz(
    ws,
    [
        "agent_user_code",
        "borc_haftasi",
        "borca_neden_olan_off_tarihleri",
        "durum",
        "odeme_haftasi",
        "odeme_tarihi",
        "odeme_tipi",
    ],
    telafi_detay_rows,
)


# 13.7 Haftalık OFF Borcu
ws = wb.create_sheet("Haftalık OFF Borcu")

sheet_yaz(
    ws,
    [
        "agent_user_code",
        "hafta",
        "off_talep_sayisi",
        "off_talep_tarihleri",
        "gercek_off_sayisi",
        "gercek_off_tarihleri",
        "standart_off_hakki",
        "telafi_borcu",
        "telafi_odeme_kapasitesi",
        "hafta_sonu_calisma_tarihleri",
    ],
    off_borc_rows,
)


# 13.8 Çift OFF
ws = wb.create_sheet("Çift OFF")

sheet_yaz(
    ws,
    [
        "agent_user_code",
        "pair_index",
        "cumartesi",
        "pazar",
        "cumartesi_durum",
        "pazar_durum",
        "cift_off_mi",
    ],
    cift_off_rows,
)

if ws.max_row >= 2:
    ws.conditional_formatting.add(
        f"G2:G{ws.max_row}",
        CellIsRule(
            operator="equal",
            formula=["1"],
            fill=fill_olustur(RENKLER["CIFT_OFF"]),
        ),
    )


# 13.9 Coverage
ws = wb.create_sheet("Coverage")

sheet_yaz(
    ws,
    [
        "tarih",
        "gun",
        "vardiya",
        "required",
        "atanan",
        "gap",
        "eksik",
        "fazla",
    ],
    coverage_rows,
)

if ws.max_row >= 2:
    ws.conditional_formatting.add(
        f"G2:G{ws.max_row}",
        CellIsRule(
            operator="greaterThan",
            formula=["0"],
            fill=fill_olustur(RENKLER["IHLAL"]),
        ),
    )


# 13.10 Kural Kontrol Özeti
ws = wb.create_sheet("Kural Kontrol Özeti")

sheet_yaz(
    ws,
    [
        "kural",
        "kontrol_kaydi",
        "ihlal_sayisi",
        "genel_sonuc",
    ],
    kural_ozet_rows,
)

if ws.max_row >= 2:
    ws.conditional_formatting.add(
        f"D2:D{ws.max_row}",
        FormulaRule(
            formula=['D2="İHLAL"'],
            fill=fill_olustur(RENKLER["IHLAL"]),
        ),
    )
    ws.conditional_formatting.add(
        f"D2:D{ws.max_row}",
        FormulaRule(
            formula=['D2="OK"'],
            fill=fill_olustur(RENKLER["OK"]),
        ),
    )


# 13.11 Kural Kontrol Detay
ws = wb.create_sheet("Kural Kontrol Detay")

sheet_yaz(
    ws,
    [
        "kural",
        "agent_user_code",
        "tarih_hafta",
        "sonuc",
        "detay",
    ],
    kural_rows,
)

if ws.max_row >= 2:
    ws.conditional_formatting.add(
        f"D2:D{ws.max_row}",
        FormulaRule(
            formula=['D2="İHLAL"'],
            fill=fill_olustur(RENKLER["IHLAL"]),
        ),
    )


# ------------------------------------------------------------
# 14) EXPORT
# ------------------------------------------------------------

wb.save(RAPOR_YOLU)

print("Excel raporu oluşturuldu:")
print(RAPOR_YOLU)
